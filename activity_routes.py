"""
routes/activity_routes.py
=========================
Handles:
  POST /api/activity/heartbeat      — receives data from desktop agent
  POST /api/activity/session-end    — end-of-day signal from agent
  GET  /api/admin/activity/summary  — admin: daily summary list
  GET  /api/admin/activity/employee/{emp_id}/detail  — per-employee detail
  GET  /api/admin/activity/export   — Excel export
  GET  /admin/activity              — admin HTML page
"""

from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Request, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from database import get_db
from models import Employee

router = APIRouter()

# ── Lazy import of activity models (added separately) ─────────────────
def _get_models():
    from activity_models import ActivityHeartbeat, ActivityAppLog, ActivityDailySummary
    return ActivityHeartbeat, ActivityAppLog, ActivityDailySummary


# ══════════════════════════════════════════════════════════════════════
# Agent Endpoints (called by desktop agent, no auth needed — LAN only)
# ══════════════════════════════════════════════════════════════════════

@router.post("/api/activity/heartbeat")
async def receive_heartbeat(request: Request, db: Session = Depends(get_db)):
    """
    Receives a 30-second heartbeat from the desktop agent.
    Upserts app_logs and daily_summary.
    """
    Heartbeat, AppLog, Summary = _get_models()
    data = await request.json()

    # Resolve employee by employee_id string → DB integer id
    emp = db.query(Employee).filter(
        Employee.employee_id == data["employee_id"]
    ).first()
    if not emp:
        return {"status": "error", "detail": "Employee not found"}

    today = date.fromisoformat(data["timestamp"][:10])
    ts    = datetime.fromisoformat(data["timestamp"])

    # ── 1. Store raw heartbeat ────────────────────────────────────────
    hb = Heartbeat(
        employee_id   = emp.id,
        timestamp     = ts,
        is_idle       = data.get("is_idle", False),
        idle_seconds  = data.get("idle_seconds", 0),
        mouse_moves   = data.get("mouse_moves", 0),
        key_presses   = data.get("key_presses", 0),
        active_app    = data.get("active_app", "Unknown"),
        os            = data.get("os", "Unknown"),
        agent_version = data.get("agent_version", ""),
    )
    db.add(hb)

    # ── 2. Upsert app time logs ───────────────────────────────────────
    app_times: dict = data.get("app_times", {})
    for app_name, seconds in app_times.items():
        if not app_name or seconds < 1:
            continue
        existing = db.query(AppLog).filter(
            AppLog.employee_id == emp.id,
            AppLog.date        == today,
            AppLog.app_name    == app_name,
        ).first()
        if existing:
            existing.seconds += int(seconds)
        else:
            db.add(AppLog(
                employee_id = emp.id,
                date        = today,
                app_name    = app_name,
                seconds     = int(seconds),
            ))

    # ── 3. Upsert daily summary ───────────────────────────────────────
    INTERVAL = 30  # agent heartbeat interval in seconds
    is_idle  = data.get("is_idle", False)

    summary = db.query(Summary).filter(
        Summary.employee_id == emp.id,
        Summary.date        == today,
    ).first()

    if not summary:
        summary = Summary(employee_id=emp.id, date=today)
        db.add(summary)

    if is_idle:
        summary.idle_seconds   += INTERVAL
    else:
        summary.active_seconds += INTERVAL

    summary.total_seconds    = summary.active_seconds + summary.idle_seconds
    summary.mouse_moves      += data.get("mouse_moves", 0)
    summary.key_presses      += data.get("key_presses", 0)
    summary.heartbeat_count  += 1
    summary.active_pct       = round(
        (summary.active_seconds / summary.total_seconds * 100)
        if summary.total_seconds > 0 else 0, 1
    )

    # Top app for today
    top = db.query(AppLog).filter(
        AppLog.employee_id == emp.id,
        AppLog.date        == today,
    ).order_by(AppLog.seconds.desc()).first()
    if top:
        summary.top_app = top.app_name

    db.commit()
    return {"status": "ok"}


@router.post("/api/activity/session-end")
async def session_end(request: Request, db: Session = Depends(get_db)):
    """Agent calls this on shutdown — just acknowledged, summary already built."""
    data = await request.json()
    print(f"[Activity] Session ended for {data.get('employee_id')} at {data.get('session_end')}")
    return {"status": "ok"}


# ══════════════════════════════════════════════════════════════════════
# Admin API — JSON data for the dashboard
# ══════════════════════════════════════════════════════════════════════

@router.get("/api/admin/activity/summary")
async def activity_summary(
    target_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    """Returns all employees' activity summary for a given date."""
    _, _, Summary = _get_models()
    day = date.fromisoformat(target_date) if target_date else date.today()

    rows = (
        db.query(Summary, Employee)
        .join(Employee, Summary.employee_id == Employee.id)
        .filter(Summary.date == day)
        .order_by(Summary.active_pct.desc())
        .all()
    )

    result = []
    for s, e in rows:
        result.append({
            "employee_id":     e.employee_id,
            "name":            e.full_name,
            "department":      e.department.name if e.department else "—",
            "active_seconds":  s.active_seconds,
            "idle_seconds":    s.idle_seconds,
            "total_seconds":   s.total_seconds,
            "active_pct":      s.active_pct,
            "mouse_moves":     s.mouse_moves,
            "key_presses":     s.key_presses,
            "top_app":         s.top_app or "—",
            "heartbeat_count": s.heartbeat_count,
        })
    return {"date": str(day), "employees": result}


@router.get("/api/admin/activity/employee/{employee_id}/detail")
async def employee_activity_detail(
    employee_id: str,
    target_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    """Returns app breakdown + hourly timeline for one employee."""
    Heartbeat, AppLog, Summary = _get_models()
    day = date.fromisoformat(target_date) if target_date else date.today()

    emp = db.query(Employee).filter(Employee.employee_id == employee_id).first()
    if not emp:
        return {"error": "Employee not found"}

    # App breakdown
    apps = (
        db.query(AppLog)
        .filter(AppLog.employee_id == emp.id, AppLog.date == day)
        .order_by(AppLog.seconds.desc())
        .all()
    )

    # Hourly timeline (group heartbeats by hour)
    heartbeats = (
        db.query(Heartbeat)
        .filter(
            Heartbeat.employee_id == emp.id,
            func.date(Heartbeat.timestamp) == day,
        )
        .order_by(Heartbeat.timestamp)
        .all()
    )

    hourly: dict = {}
    for hb in heartbeats:
        hour = hb.timestamp.hour
        if hour not in hourly:
            hourly[hour] = {"active": 0, "idle": 0}
        if hb.is_idle:
            hourly[hour]["idle"] += 30
        else:
            hourly[hour]["active"] += 30

    return {
        "employee_id": employee_id,
        "name":        emp.full_name,
        "date":        str(day),
        "app_breakdown": [
            {"app": a.app_name, "seconds": a.seconds,
             "minutes": round(a.seconds / 60, 1)}
            for a in apps
        ],
        "hourly_timeline": [
            {"hour": h, "active": v["active"], "idle": v["idle"]}
            for h, v in sorted(hourly.items())
        ],
    }


# ══════════════════════════════════════════════════════════════════════
# Excel Export
# ══════════════════════════════════════════════════════════════════════

@router.get("/api/admin/activity/export")
async def export_activity_excel(
    target_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    """Export daily activity report as .xlsx"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _, AppLog, Summary = _get_models()
    day = date.fromisoformat(target_date) if target_date else date.today()

    rows = (
        db.query(Summary, Employee)
        .join(Employee, Summary.employee_id == Employee.id)
        .filter(Summary.date == day)
        .order_by(Summary.active_pct.desc())
        .all()
    )

    wb = Workbook()

    # ── Sheet 1: Daily Summary ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"Summary {day}"

    hdr_fill   = PatternFill("solid", start_color="0F172A")
    hdr_font   = Font(color="FFFFFF", bold=True, name="Arial", size=11)
    green_fill = PatternFill("solid", start_color="DCFCE7")
    red_fill   = PatternFill("solid", start_color="FEE2E2")
    border     = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
    )

    # Title row
    ws1.merge_cells("A1:K1")
    ws1["A1"] = f"Employee Activity Report — {day.strftime('%d %B %Y')}"
    ws1["A1"].font = Font(bold=True, size=14, name="Arial", color="0F172A")
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 28

    headers = [
        "Employee ID", "Name", "Department",
        "Active Time", "Idle Time", "Total Time",
        "Active %", "Mouse Moves", "Keystrokes",
        "Top App", "Heartbeats"
    ]
    ws1.append([])  # blank row
    ws1.append(headers)
    hdr_row = 3
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=hdr_row, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    def fmt_time(seconds: int) -> str:
        h, m = divmod(seconds // 60, 60)
        return f"{h}h {m:02d}m"

    for s, e in rows:
        row = [
            e.employee_id,
            e.full_name,
            e.department.name if e.department else "—",
            fmt_time(s.active_seconds),
            fmt_time(s.idle_seconds),
            fmt_time(s.total_seconds),
            f"{s.active_pct:.1f}%",
            s.mouse_moves,
            s.key_presses,
            s.top_app or "—",
            s.heartbeat_count,
        ]
        ws1.append(row)
        r = ws1.max_row
        fill = green_fill if s.active_pct >= 60 else red_fill
        for col in range(1, 12):
            cell = ws1.cell(row=r, column=col)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if col == 7:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 22, 16, 12, 12, 12, 10, 13, 12, 22, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: App Breakdown ────────────────────────────────────────
    ws2 = wb.create_sheet("App Breakdown")
    ws2.append(["Employee ID", "Name", "App Name", "Minutes", "Hours"])
    for col in range(1, 6):
        cell = ws2.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for s, e in rows:
        apps = (
            db.query(AppLog)
            .filter(AppLog.employee_id == e.id, AppLog.date == day)
            .order_by(AppLog.seconds.desc())
            .all()
        )
        for a in apps:
            ws2.append([
                e.employee_id,
                e.full_name,
                a.app_name,
                round(a.seconds / 60, 1),
                round(a.seconds / 3600, 2),
            ])

    for col, w in zip("ABCDE", [12, 22, 25, 10, 8]):
        ws2.column_dimensions[col].width = w

    # ── Stream response ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"activity_report_{day}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ══════════════════════════════════════════════════════════════════════
# Admin HTML Page
# ══════════════════════════════════════════════════════════════════════

@router.get("/admin/activity")
async def activity_page(request: Request, db: Session = Depends(get_db)):
    from fastapi.templating import Jinja2Templates
    templates = Jinja2Templates(directory="templates")

    admin_id = request.session.get("admin_id")
    if not admin_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/login")

    today = date.today()
    return templates.TemplateResponse("admin/activity.html", {
        "request":    request,
        "today":      today,
        "admin_name": request.session.get("admin_name", "Admin"),
    })