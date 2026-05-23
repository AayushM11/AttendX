import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Request, Form
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, text
from database import get_db
from models import Attendance, Employee, Holiday, LeaveRequest, LeaveType, Department, FaceEmbedding
from recognition import recognize_face, generate_embeddings_for_employee, check_duplicate_face, load_image_from_bytes
from geofence import check_geofence, get_company_settings
import datetime
import json
import io
import shutil
import uuid
import random
import string

router = APIRouter()

# OTP STORE

_otp_store: dict = {}
OTP_EXPIRY_MINUTES = 10


# EMAIL HELPER

async def _send_otp_email(to_email: str, otp: str, first_name: str) -> bool:
    username = os.getenv("MAIL_USERNAME")
    password = os.getenv("MAIL_PASSWORD")

    if not username or not password:
        print("\n" + "=" * 50)
        print("  [OTP - DEV MODE] Email not configured")
        print(f"  To      : {to_email}")
        print(f"  OTP     : {otp}")
        print(f"  Expires : {OTP_EXPIRY_MINUTES} minutes")
        print("=" * 50 + "\n")
        return False

    try:
        from fastapi_mail import FastMail, MessageSchema, MessageType, ConnectionConfig
        conf = ConnectionConfig(
            MAIL_USERNAME   = username,
            MAIL_PASSWORD   = password,
            MAIL_FROM       = os.getenv("MAIL_FROM", username),
            MAIL_PORT       = int(os.getenv("MAIL_PORT", "587")),
            MAIL_SERVER     = os.getenv("MAIL_SERVER", "smtp.gmail.com"),
            MAIL_STARTTLS   = True,
            MAIL_SSL_TLS    = False,
            USE_CREDENTIALS = True,
            VALIDATE_CERTS  = True,
        )
        html_body = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;margin:auto;
                    background:#0F172A;color:white;border-radius:16px;padding:32px;">
            <div style="text-align:center;margin-bottom:24px;">
                <div style="background:linear-gradient(135deg,#2563EB,#06B6D4);
                            width:60px;height:60px;border-radius:14px;
                            margin:auto;line-height:60px;font-size:28px;">🔐</div>
                <h2 style="margin-top:16px;color:white;">AttendX Verification</h2>
            </div>
            <p style="color:#94A3B8;">Hi {first_name},</p>
            <p style="color:#94A3B8;">Your one-time password for AttendX registration is:</p>
            <div style="background:#1E293B;border:1px solid #334155;border-radius:12px;
                        padding:24px;text-align:center;margin:24px 0;">
                <div style="font-size:42px;font-weight:900;letter-spacing:12px;color:#60A5FA;">{otp}</div>
                <div style="color:#64748B;font-size:13px;margin-top:8px;">Expires in {OTP_EXPIRY_MINUTES} minutes</div>
            </div>
            <p style="color:#64748B;font-size:12px;text-align:center;">
                Do not share this OTP with anyone.
            </p>
        </div>
        """
        message = MessageSchema(
            subject    = f"Your AttendX OTP: {otp}",
            recipients = [to_email],
            body       = html_body,
            subtype    = MessageType.html,
        )
        fm = FastMail(conf)
        await fm.send_message(message)
        print(f"[OTP] ✅ Email sent to {to_email}")
        return True
    except Exception as e:
        print(f"[OTP] ❌ Email failed: {e}")
        print(f"[OTP] FALLBACK OTP for {to_email}: {otp}")
        return False


# OTP ENDPOINTS

@router.post("/api/register/send-otp")
async def send_registration_otp(request: Request, db: Session = Depends(get_db)):
    data       = await request.json()
    email      = data.get("email", "").strip().lower()
    first_name = data.get("first_name", "there").strip()

    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="A valid email address is required.")
    if db.query(Employee).filter(Employee.email == email).first():
        raise HTTPException(status_code=400, detail="This email is already registered. Please login instead.")

    otp = ''.join(random.choices(string.digits, k=6))
    _otp_store[email] = {
        "otp":      otp,
        "expires":  datetime.datetime.now() + datetime.timedelta(minutes=OTP_EXPIRY_MINUTES),
        "verified": False,
    }
    sent = await _send_otp_email(email, otp, first_name)
    return {
        "success":            True,
        "sent":               sent,
        "message":            f"OTP sent to {email}." if sent else "Check server terminal for OTP.",
        "expires_in_minutes": OTP_EXPIRY_MINUTES,
    }


@router.post("/api/register/verify-otp")
async def verify_registration_otp(request: Request):
    data  = await request.json()
    email = data.get("email", "").strip().lower()
    otp   = data.get("otp", "").strip()
    entry = _otp_store.get(email)

    if not entry:
        raise HTTPException(status_code=400, detail="No OTP found. Please request a new one.")
    if datetime.datetime.now() > entry["expires"]:
        del _otp_store[email]
        raise HTTPException(status_code=400, detail="OTP expired. Please request a new one.")
    if entry["otp"] != otp:
        raise HTTPException(status_code=400, detail="Incorrect OTP. Please try again.")

    _otp_store[email]["verified"] = True
    return {"success": True, "message": "Email verified. Proceed with registration."}


# REGISTRATION

@router.post("/api/register")
async def self_register(
    first_name:      str              = Form(...),
    last_name:       str              = Form(...),
    email:           str              = Form(...),
    phone:           str              = Form(...),
    department_id:   int              = Form(...),
    designation:     str              = Form(...),
    date_of_joining: str              = Form(...),
    face_photos:     list[UploadFile] = File(...),
    db:              Session          = Depends(get_db),
):
    email = email.strip().lower()

    otp_entry = _otp_store.get(email)
    if not otp_entry or not otp_entry.get("verified"):
        raise HTTPException(status_code=400, detail="Email not verified. Complete OTP verification first.")

    if db.query(Employee).filter(Employee.email == email).first():
        raise HTTPException(status_code=400, detail=f"Email '{email}' already registered.")

    if db.query(Employee).filter(
        Employee.first_name == first_name,
        Employee.last_name  == last_name,
        Employee.is_active  == True
    ).first():
        raise HTTPException(status_code=400, detail=f"Employee '{first_name} {last_name}' already exists.")

    valid_photos = [p for p in face_photos if p and p.filename]
    if not valid_photos:
        raise HTTPException(status_code=400, detail="At least 1 face photo is required.")

    photo_bytes_list = [await photo.read() for photo in valid_photos]

    rgb_images = [load_image_from_bytes(b) for b in photo_bytes_list]
    rgb_images = [r for r in rgb_images if r is not None]

    if rgb_images:
        print(f"[Register] Checking {len(rgb_images)} photo(s) for duplicate face...")
        is_dup, dup_id, dup_name = check_duplicate_face(rgb_images, db)
        if is_dup:
            raise HTTPException(status_code=400, detail=(
                f"Your face is already registered as '{dup_name}' (ID: {dup_id}). "
                f"Please login with Employee ID: {dup_id}."
            ))
        print("[Register] ✅ No duplicate face.")

    emp_count   = db.query(Employee).count() + 1
    employee_id = f"EMP{emp_count:04d}"
    base_dir    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    upload_dir  = os.path.join(base_dir, "uploaded_images", employee_id)
    os.makedirs(upload_dir, exist_ok=True)

    image_path = None
    for i, (photo, photo_bytes) in enumerate(zip(valid_photos, photo_bytes_list)):
        ext      = (photo.filename or "jpg").rsplit(".", 1)[-1].lower()
        filename = f"face_{i+1:02d}.{ext}"
        path     = os.path.join(upload_dir, filename)
        with open(path, "wb") as f:
            f.write(photo_bytes)
        if i == 0:
            image_path = f"uploaded_images/{employee_id}/{filename}"

    emp = Employee(
        employee_id     = employee_id,
        first_name      = first_name,
        last_name       = last_name,
        email           = email,
        phone           = phone,
        department_id   = department_id,
        designation     = designation,
        date_of_joining = datetime.date.fromisoformat(date_of_joining),
        profile_image   = image_path,
        is_active       = True,
        self_registered = True,
    )
    db.add(emp)
    db.commit()
    db.refresh(emp)
    _otp_store.pop(email, None)

    embedding_error  = None
    embeddings_count = 0
    try:
        result           = generate_embeddings_for_employee(emp.id, db)
        embeddings_count = result.get("embeddings_count", 0)
        if not result.get("success"):
            embedding_error = result.get("message")
    except Exception as e:
        embedding_error = str(e)

    print(f"[Registration] {employee_id} created. Embeddings: {embeddings_count}.")
    return {
        "success":              True,
        "employee_id":          employee_id,
        "name":                 emp.full_name,
        "message":              f"Registration successful! Your Employee ID is {employee_id}",
        "embeddings_generated": embeddings_count,
        "embedding_error":      embedding_error,
    }


@router.post("/api/check-email")
async def check_email(request: Request, db: Session = Depends(get_db)):
    data   = await request.json()
    email  = data.get("email", "").strip().lower()
    exists = db.query(Employee).filter(Employee.email == email).first() is not None
    return {"exists": exists}


# REGENERATE / CHECK EMBEDDINGS

@router.get("/api/regenerate-embeddings/{employee_code}")
async def regenerate_embeddings(employee_code: str, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(
        Employee.employee_id == employee_code, Employee.is_active == True
    ).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    try:
        result = generate_embeddings_for_employee(employee.id, db)
        return {
            "success":          result.get("success", False),
            "employee_id":      employee_code,
            "name":             employee.full_name,
            "embeddings_count": result.get("embeddings_count", 0),
            "message":          result.get("message", ""),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/check-embeddings/{employee_code}")
async def check_embeddings(employee_code: str, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.employee_id == employee_code).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    from models import FaceEmbedding

    count      = db.query(FaceEmbedding).filter(FaceEmbedding.employee_id == employee.id).count()
    base_dir   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    upload_dir = os.path.join(base_dir, "uploaded_images", employee_code)
    photos     = []
    if os.path.exists(upload_dir):
        photos = [f for f in os.listdir(upload_dir) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    return {
        "employee_id":      employee_code,
        "name":             employee.full_name,
        "embeddings_in_db": count,
        "photos_on_disk":   len(photos),
        "photo_files":      photos,
        "status":           "ok" if count > 0 else "no_embeddings",
    }


# GEOFENCE SETTINGS API (used by admin panel + app)

@router.get("/api/geofence-settings")
async def get_geofence_settings(db: Session = Depends(get_db)):
    """Returns current geofence settings. Used by Flutter app to show map."""
    settings = get_company_settings(db)
    return settings


@router.post("/api/geofence-settings")
async def update_geofence_settings(request: Request, db: Session = Depends(get_db)):
    """Update geofence settings from admin panel."""
    data = await request.json()
    updates = {
        "company_lat":  str(data.get("lat",       "22.55630805480531")),
        "company_lng":  str(data.get("lng",       "72.95262732498357")),
        "geo_radius_m": str(data.get("radius_m",  "100")),
        "geo_enabled":  str(data.get("enabled",   "true")).lower(),
        "company_name": str(data.get("name",      "Company HQ")),
    }
    for key, value in updates.items():
        existing = db.execute(
            text("SELECT id FROM company_settings WHERE key = :k"), {"k": key}
        ).fetchone()
        if existing:
            db.execute(
                text("UPDATE company_settings SET value=:v, updated_at=NOW() WHERE key=:k"),
                {"v": value, "k": key}
            )
        else:
            db.execute(
                text("INSERT INTO company_settings (key, value) VALUES (:k, :v)"),
                {"k": key, "v": value}
            )
    db.commit()
    return {"success": True, "message": "Geofence settings updated."}


# FACE RECOGNITION ATTENDANCE  (with geofence)

@router.post("/api/recognize")
async def recognize_and_mark(
    file:                 UploadFile = File(...),
    latitude:             float      = Form(None),
    longitude:            float      = Form(None),
    expected_employee_id: str        = Form(None),
    db:                   Session    = Depends(get_db),
):
    # ── Step 1: Geofence check BEFORE face recognition 
    # This saves processing time — no point recognising face if outside office
    settings = get_company_settings(db)

    if settings["enabled"]:
        if latitude is None or longitude is None:
            return JSONResponse({
                "status":   "location_required",
                "message":  "Location permission is required for attendance.",
                "distance": None,
                "radius":   settings["radius_m"],
            })

        is_inside, distance, _ = check_geofence(latitude, longitude, db)

        print(f"[Geofence] Employee at ({latitude:.5f}, {longitude:.5f}) "
              f"— {distance}m from office (radius={settings['radius_m']}m) "
              f"— {'✅ inside' if is_inside else '❌ outside'}")

        if not is_inside:
            return JSONResponse({
                "status":        "outside_geofence",
                "message":       (
                    f"You are {distance:.0f}m away from {settings['name']}. "
                    f"Please move within {settings['radius_m']:.0f}m to mark attendance."
                ),
                "distance":      distance,
                "radius":        settings["radius_m"],
                "company_name":  settings["name"],
            })
    else:
        distance = None

    #  Step 2: Face recognition
    image_bytes = await file.read()
    employee_id, employee_name, status = recognize_face(image_bytes, db)

    if status == "no_face":
        return JSONResponse({"status": "no_face"})
    if status in ("unknown", "error"):
        return JSONResponse({"status": "unknown"})

    #  Step 3: Security: verify face belongs to logged-in employee 
    if expected_employee_id:
        expected_emp = db.query(Employee).filter(
            Employee.employee_id == expected_employee_id,
            Employee.is_active == True
        ).first()
        if expected_emp and expected_emp.id != employee_id:
            return JSONResponse({
                "status":           "wrong_person",
                "recognized_name":  employee_name,
                "expected_name":    expected_emp.full_name,
            })

    today = datetime.date.today()
    now   = datetime.datetime.now()

    record = db.query(Attendance).filter(
        and_(Attendance.employee_id == employee_id, Attendance.date == today)
    ).first()

    if not record:
        record = Attendance(
            employee_id       = employee_id,
            date              = today,
            check_in          = now,
            status            = "present",
            latitude          = latitude,
            longitude         = longitude,
            distance_meters   = distance,
            location_verified = True if settings["enabled"] else None,
        )
        db.add(record)
        db.commit()
        return JSONResponse({
            "status":   "check_in",
            "name":     employee_name,
            "time":     now.strftime("%I:%M %p"),
            "date":     today.strftime("%d %B %Y"),
            "distance": distance,
        })
    else:
        if record.check_out is None:
            record.check_out          = now
            record.location_verified  = True if settings["enabled"] else None
            if record.check_in:
                delta             = now - record.check_in
                record.work_hours = round(delta.total_seconds() / 3600, 2)
            db.commit()
            return JSONResponse({
                "status":     "check_out",
                "name":       employee_name,
                "time":       now.strftime("%I:%M %p"),
                "work_hours": record.work_hours,
                "date":       today.strftime("%d %B %Y"),
                "distance":   distance,
            })
        else:
            return JSONResponse({
                "status":     "already_checked_out",
                "name":       employee_name,
                "check_in":   record.check_in.strftime("%I:%M %p"),
                "check_out":  record.check_out.strftime("%I:%M %p"),
                "work_hours": record.work_hours,
            })


# EMPLOYEE INFO & STATS

@router.get("/api/employee/{employee_code}")
async def get_employee_info(employee_code: str, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(
        Employee.employee_id == employee_code, Employee.is_active == True
    ).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found or deactivated")

    today               = datetime.date.today()
    current_month_start = today.replace(day=1)
    month_records       = db.query(Attendance).filter(
        and_(Attendance.employee_id == employee.id,
             Attendance.date >= current_month_start,
             Attendance.date <= today)
    ).all()

    present_days     = sum(1 for r in month_records if r.status == "present")
    total_work_hours = sum(r.work_hours or 0 for r in month_records)
    today_record     = next((r for r in month_records if r.date == today), None)

    return {
        "id":              employee.id,
        "employee_id":     employee.employee_id,
        "name":            employee.full_name,
        "email":           employee.email,
        "phone":           employee.phone,
        "department":      employee.department.name if employee.department else None,
        "designation":     employee.designation,
        "date_of_joining": str(employee.date_of_joining) if employee.date_of_joining else None,
        "profile_image":   employee.profile_image,
        "today_status": {
            "checked_in":  today_record.check_in.strftime("%I:%M %p")  if today_record and today_record.check_in  else None,
            "checked_out": today_record.check_out.strftime("%I:%M %p") if today_record and today_record.check_out else None,
            "work_hours":  today_record.work_hours if today_record else 0,
        },
        "monthly_stats": {
            "present_days":     present_days,
            "total_work_hours": round(total_work_hours, 1),
            "month":            today.strftime("%B %Y"),
        },
    }


@router.get("/api/employee/{employee_code}/attendance")
async def get_employee_attendance(
    employee_code: str, month: int = None, year: int = None,
    db: Session = Depends(get_db),
):
    employee = db.query(Employee).filter(
        Employee.employee_id == employee_code, Employee.is_active == True
    ).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    today  = datetime.date.today()
    month  = month or today.month
    year   = year  or today.year
    start  = datetime.date(year, month, 1)
    end    = (datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
              if month < 12 else datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1))

    records        = db.query(Attendance).filter(
        and_(Attendance.employee_id == employee.id,
             Attendance.date >= start, Attendance.date <= end)
    ).order_by(Attendance.date.desc()).all()

    holidays       = db.query(Holiday).filter(and_(Holiday.date >= start, Holiday.date <= end)).all()
    holiday_dates  = {h.date: h.name for h in holidays}
    leave_requests = db.query(LeaveRequest).filter(
        and_(LeaveRequest.employee_id == employee.id,
             LeaveRequest.status == "approved",
             LeaveRequest.start_date <= end, LeaveRequest.end_date >= start)
    ).all()

    attendance_map = {r.date: r for r in records}
    calendar_data  = []
    current        = start

    while current <= min(end, today):
        day_info = {
            "date": str(current), "day": current.strftime("%A"),
            "is_holiday": current in holiday_dates,
            "holiday_name": holiday_dates.get(current),
            "is_weekend": current.weekday() >= 5,
        }
        if current in attendance_map:
            rec = attendance_map[current]
            day_info.update({
                "status":     rec.status,
                "check_in":   rec.check_in.strftime("%I:%M %p")  if rec.check_in  else None,
                "check_out":  rec.check_out.strftime("%I:%M %p") if rec.check_out else None,
                "work_hours": rec.work_hours or 0,
            })
        elif current in holiday_dates or current.weekday() >= 5:
            day_info["status"] = "holiday" if current in holiday_dates else "weekend"
        else:
            on_leave           = any(lr.start_date <= current <= lr.end_date for lr in leave_requests)
            day_info["status"] = "leave" if on_leave else "absent"
        calendar_data.append(day_info)
        current += datetime.timedelta(days=1)

    return {"attendance": calendar_data, "month": start.strftime("%B %Y")}


# LEAVES

@router.get("/api/leave-types")
async def get_leave_types(db: Session = Depends(get_db)):
    types = db.query(LeaveType).all()
    return [{"id": t.id, "name": t.name, "max_days": t.max_days} for t in types]


@router.post("/api/apply-leave")
async def apply_leave(request: Request, db: Session = Depends(get_db)):
    data     = await request.json()
    employee = db.query(Employee).filter(Employee.employee_id == data.get("employee_id")).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    leave = LeaveRequest(
        employee_id   = employee.id,
        leave_type_id = data["leave_type_id"],
        start_date    = datetime.date.fromisoformat(data["start_date"]),
        end_date      = datetime.date.fromisoformat(data["end_date"]),
        reason        = data.get("reason", ""),
    )
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return {"success": True, "leave_id": leave.id, "status": "pending"}


@router.get("/api/employee/{employee_code}/leaves")
async def get_employee_leaves(employee_code: str, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.employee_id == employee_code).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == employee.id
    ).order_by(LeaveRequest.applied_at.desc()).all()
    return [{
        "id": l.id, "type": l.leave_type.name if l.leave_type else "Unknown",
        "start_date": str(l.start_date), "end_date": str(l.end_date),
        "reason": l.reason, "status": l.status,
        "admin_comment": l.admin_comment, "applied_at": str(l.applied_at),
    } for l in leaves]


# HOLIDAYS

@router.get("/api/holidays")
async def get_holidays(year: int = None, db: Session = Depends(get_db)):
    year     = year or datetime.date.today().year
    holidays = db.query(Holiday).filter(
        func.extract("year", Holiday.date) == year
    ).order_by(Holiday.date).all()
    return [{"name": h.name, "date": str(h.date), "description": h.description} for h in holidays]


# ADMIN REPORTS

@router.get("/api/admin/attendance/download/daily")
async def download_daily_excel(date: str = None, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    target_date = datetime.date.fromisoformat(date) if date else datetime.date.today()
    # Block future date downloads
    if target_date > datetime.date.today():
        raise HTTPException(
            status_code=400,
            detail=f"Cannot download attendance for future date {target_date}. Only past or today's dates are allowed."
        )
    records     = db.query(Attendance).filter(Attendance.date == target_date).all()
    wb = Workbook(); ws = wb.active; ws.title = f"Attendance {target_date}"
    header_fill = PatternFill("solid", start_color="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    headers = ["#", "Employee ID", "Name", "Department", "Designation",
               "Check In", "Check Out", "Work Hours", "Status", "Distance(m)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    for i, r in enumerate(records, 1):
        emp  = r.employee
        _dist = getattr(r, "distance_meters", None)
        _dist_str = (str(int(_dist)) + "m") if _dist is not None else "—"
        row_data = [
            i, emp.employee_id, emp.full_name,
            emp.department.name if emp.department else "—", emp.designation or "—",
            r.check_in.strftime("%I:%M %p")  if r.check_in  else "—",
            r.check_out.strftime("%I:%M %p") if r.check_out else "—",
            r.work_hours or 0, r.status.upper(),
            _dist_str,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = border; cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0: cell.fill = PatternFill("solid", start_color="EBF0F7")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{target_date}.xlsx"})


@router.get("/api/admin/attendance/download/monthly")
async def download_monthly_excel(month: int = None, year: int = None, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = datetime.date.today(); month = month or today.month; year = year or today.year
    # Block future month downloads
    requested_month_start = datetime.date(year, month, 1)
    if requested_month_start > today:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot download attendance for future month {requested_month_start.strftime('%B %Y')}."
        )
    start = datetime.date(year, month, 1)
    end   = (datetime.date(year, month+1, 1) - datetime.timedelta(days=1)
             if month < 12 else datetime.date(year+1, 1, 1) - datetime.timedelta(days=1))
    employees = db.query(Employee).filter(Employee.is_active == True).all()
    wb = Workbook(); ws = wb.active; ws.title = start.strftime("%B %Y")
    header_fill = PatternFill("solid", start_color="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    headers = ["#", "Employee ID", "Name", "Department", "Present Days", "Absent Days", "Total Work Hours"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = border
    for i, emp in enumerate(employees, 1):
        records     = db.query(Attendance).filter(
            and_(Attendance.employee_id == emp.id,
                 Attendance.date >= start, Attendance.date <= end)).all()
        present     = sum(1 for r in records if r.status == "present")
        total_hours = sum(r.work_hours or 0 for r in records)
        row_data    = [i, emp.employee_id, emp.full_name,
                       emp.department.name if emp.department else "—",
                       present, 0, round(total_hours, 1)]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.border = border; cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0: cell.fill = PatternFill("solid", start_color="EBF0F7")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=monthly_{start.strftime('%B_%Y')}.xlsx"})


@router.get("/api/threshold-test/{employee_code}")
async def threshold_test(employee_code: str, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(
        Employee.employee_id == employee_code, Employee.is_active == True
    ).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    embeddings = db.query(FaceEmbedding).filter(FaceEmbedding.employee_id == employee.id).all()
    return {
        "employee_id": employee_code, "name": employee.full_name,
        "embeddings_stored": len(embeddings), "current_threshold": 0.40,
        "status": "ready" if embeddings else "no_embeddings — register first",
    }
